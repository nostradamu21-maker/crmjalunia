"""
app.py - Jalunia CRM Backend
Flask API + serves React frontend + Campaign auto-send engine
"""
import os
import json
import smtplib
import imaplib
import email as emaillib
import ssl
import re
import secrets
import time
import random
from datetime import datetime, timezone, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from functools import wraps

import hmac
import hashlib
import base64
from flask import Flask, request, jsonify, send_from_directory, abort
from flask_cors import CORS
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from models import db, Prospect, EmailLog, Setting, CampaignRun

# --- App Config ---------------------------------------------------------------
JWT_EXPIRATION_HOURS = int(os.environ.get("JWT_EXPIRATION_HOURS", "24"))
ALLOWED_ORIGINS = os.environ.get("ALLOWED_ORIGINS", "https://crmjalunia.onrender.com").split(",")

def create_app():
    app = Flask(__name__, static_folder="static", static_url_path="")

    database_url = os.environ.get("DATABASE_URL", "sqlite:////tmp/jalunia_crm.db")
    if database_url.startswith("postgres://"):
        database_url = database_url.replace("postgres://", "postgresql://", 1)

    app.config["SQLALCHEMY_DATABASE_URI"] = database_url
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
    app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", secrets.token_hex(32))

    db.init_app(app)
    CORS(app, origins=ALLOWED_ORIGINS)

    with app.app_context():
        db.create_all()
        # Auto-migrate: add new columns to existing tables
        _auto_migrate(app)

    return app

def _auto_migrate(app):
    """Add missing columns to existing tables (safe to run multiple times)."""
    migrations = [
        ("prospects", "bounce_count", "INTEGER DEFAULT 0"),
        ("prospects", "score", "INTEGER DEFAULT 0"),
        ("prospects", "email_opened", "BOOLEAN DEFAULT false"),
        ("prospects", "unsubscribe_token", "VARCHAR(64)"),
        ("email_logs", "error_message", "TEXT"),
        ("email_logs", "tracking_id", "VARCHAR(64)"),
        ("email_logs", "opened_at", "TIMESTAMP"),
        ("email_logs", "open_count", "INTEGER DEFAULT 0"),
    ]
    for table, column, col_type in migrations:
        try:
            with db.engine.connect() as conn:
                conn.execute(db.text(f"ALTER TABLE {table} ADD COLUMN IF NOT EXISTS {column} {col_type}"))
                conn.commit()
        except Exception:
            pass  # Column already exists or other non-critical error

app = create_app()
limiter = Limiter(get_remote_address, app=app, default_limits=["200 per minute"])

# Backfill unsubscribe tokens for existing prospects (batched)
with app.app_context():
    try:
        batch = Prospect.query.filter(Prospect.unsubscribe_token.is_(None)).limit(500).all()
        for p in batch:
            p.unsubscribe_token = secrets.token_urlsafe(32)
        if batch:
            db.session.commit()
    except Exception:
        db.session.rollback()

# 1x1 transparent GIF for tracking pixel
TRACKING_GIF = b'GIF89a\x01\x00\x01\x00\x80\x00\x00\xff\xff\xff\x00\x00\x00!\xf9\x04\x00\x00\x00\x00\x00,\x00\x00\x00\x00\x01\x00\x01\x00\x00\x02\x02D\x01\x00;'
BASE_URL = os.environ.get("BASE_URL", "https://crmjalunia.onrender.com")

# --- Auth (JWT) ---------------------------------------------------------------
CRM_PASSWORD = os.environ.get("CRM_PASSWORD", "jalunia2026")
API_SECRET = os.environ.get("API_SECRET", "")

SKIP_STATUSES = ("replied", "meeting", "converted", "not_interested", "unsubscribed", "bounced")

def _b64url_encode(data):
    return base64.urlsafe_b64encode(data).rstrip(b"=").decode()

def _b64url_decode(s):
    s += "=" * (4 - len(s) % 4)
    return base64.urlsafe_b64decode(s)

def _create_token():
    header = _b64url_encode(json.dumps({"alg": "HS256", "typ": "JWT"}).encode())
    payload = _b64url_encode(json.dumps({
        "sub": "crm_user",
        "iat": int(datetime.now(timezone.utc).timestamp()),
        "exp": int((datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)).timestamp()),
    }).encode())
    sig_input = f"{header}.{payload}".encode()
    signature = _b64url_encode(hmac.new(app.config["SECRET_KEY"].encode(), sig_input, hashlib.sha256).digest())
    return f"{header}.{payload}.{signature}"

def _verify_token(token):
    try:
        parts = token.split(".")
        if len(parts) != 3:
            return False
        sig_input = f"{parts[0]}.{parts[1]}".encode()
        expected_sig = _b64url_encode(hmac.new(app.config["SECRET_KEY"].encode(), sig_input, hashlib.sha256).digest())
        if not hmac.compare_digest(parts[2], expected_sig):
            return False
        payload = json.loads(_b64url_decode(parts[1]))
        if payload.get("exp", 0) < datetime.now(timezone.utc).timestamp():
            return False
        return True
    except Exception:
        return False

def require_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.headers.get("Authorization", "")
        token = auth.replace("Bearer ", "") if auth.startswith("Bearer ") else ""
        if not token or not _verify_token(token):
            return jsonify({"error": "Non autorise"}), 401
        return f(*args, **kwargs)
    return decorated

def require_auth_or_secret(f):
    """Auth via JWT OR via ?secret= query param (for cron jobs)."""
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.headers.get("Authorization", "")
        token = auth.replace("Bearer ", "") if auth.startswith("Bearer ") else ""
        if token and _verify_token(token):
            return f(*args, **kwargs)
        secret = request.args.get("secret", "")
        if API_SECRET and secret and hmac.compare_digest(secret, API_SECRET):
            return f(*args, **kwargs)
        return jsonify({"error": "Non autorise"}), 401
    return decorated

# --- Template Engine ----------------------------------------------------------
def _render_template(text, prospect):
    """Replace {nom}, {ville}, {type}, {region}, {site} in email templates."""
    if not text:
        return text
    replacements = {
        "{nom}": prospect.nom or "",
        "{ville}": prospect.ville or "",
        "{type}": prospect.type or "",
        "{region}": prospect.region or prospect.ville or "",
        "{site}": prospect.site_web or "",
        "{email}": prospect.email or "",
        "{telephone}": prospect.telephone or "",
    }
    for key, val in replacements.items():
        text = text.replace(key, val)
    return text

def _is_valid_email(email):
    return bool(email and re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email.strip()))

def _calculate_score(prospect):
    """Auto-score a prospect based on engagement signals."""
    s = 0
    if prospect.email and _is_valid_email(prospect.email):
        s += 10
    if prospect.site_web:
        s += 5
    if prospect.telephone:
        s += 5
    if prospect.note_google and prospect.note_google >= 4.0:
        s += 10
    if prospect.note_google and prospect.note_google >= 4.5:
        s += 5
    if prospect.email_opened:
        s += 20
    if prospect.status == "replied":
        s += 50
    if prospect.status == "meeting":
        s += 70
    if prospect.status == "converted":
        s += 100
    if prospect.status in ("bounced", "unsubscribed", "not_interested"):
        s = max(s - 50, 0)
    prospect.score = s
    return s

def _get_france_hour():
    """Get current hour in France (CET/CEST)."""
    now = datetime.now(timezone.utc)
    month = now.month
    # CEST: last Sunday of March to last Sunday of October = UTC+2
    # CET: rest of year = UTC+1
    offset = 2 if 4 <= month <= 9 else 1
    if month == 3 and now.day >= 25:
        offset = 2
    if month == 10 and now.day < 25:
        offset = 2
    return (now + timedelta(hours=offset)).hour

# --- SMTP Helpers -------------------------------------------------------------
def _get_smtp_config():
    return {
        "host": Setting.get("smtp_host", ""),
        "port": int(Setting.get("smtp_port", "587")),
        "user": Setting.get("smtp_user", ""),
        "password": Setting.get("smtp_pass", ""),
        "sender_email": Setting.get("sender_email", "") or Setting.get("smtp_user", ""),
        "sender_name": Setting.get("sender_name", "Jalunia"),
    }

def _open_smtp_connection():
    """Open a reusable SMTP connection."""
    cfg = _get_smtp_config()
    if not cfg["host"] or not cfg["user"]:
        return None
    context = ssl.create_default_context()
    server = smtplib.SMTP(cfg["host"], cfg["port"], timeout=30)
    server.starttls(context=context)
    server.login(cfg["user"], cfg["password"])
    return server

def _send_one_email(prospect, subject, body, email_num, smtp_server=None, seen_emails=None):
    """Send one email with tracking pixel + unsubscribe link. Returns (success, error_message)."""
    # Duplicate detection
    if seen_emails is not None:
        email_lower = prospect.email.strip().lower()
        if email_lower in seen_emails:
            return False, "duplicate"
        seen_emails.add(email_lower)

    subject = _render_template(subject, prospect)
    body = _render_template(body, prospect)
    cfg = _get_smtp_config()

    # Generate tracking ID
    tracking_id = secrets.token_urlsafe(32)

    # Ensure prospect has unsubscribe token
    if not prospect.unsubscribe_token:
        prospect.unsubscribe_token = secrets.token_urlsafe(32)

    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"] = f"{cfg['sender_name']} <{cfg['sender_email']}>"
        msg["To"] = prospect.email
        msg["Reply-To"] = cfg["sender_email"]
        # List-Unsubscribe header for Gmail/Outlook one-click unsubscribe
        msg["List-Unsubscribe"] = f"<{BASE_URL}/api/unsubscribe/{prospect.unsubscribe_token}>"
        msg["List-Unsubscribe-Post"] = "List-Unsubscribe=One-Click"

        msg.attach(MIMEText(body, "plain", "utf-8"))

        html_body = body.replace("\n", "<br>")
        # Add tracking pixel + unsubscribe footer
        unsub_link = f"{BASE_URL}/api/unsubscribe/{prospect.unsubscribe_token}"
        pixel_url = f"{BASE_URL}/api/track/{tracking_id}"
        html_full = (
            f"<html><body style='font-family:Arial,sans-serif;font-size:14px;color:#333;line-height:1.6'>"
            f"{html_body}"
            f"<p style='margin-top:30px;padding-top:10px;border-top:1px solid #eee;font-size:11px;color:#999;'>"
            f"Si vous ne souhaitez plus recevoir nos emails, "
            f"<a href='{unsub_link}' style='color:#999;'>cliquez ici</a>.</p>"
            f"<img src='{pixel_url}' width='1' height='1' style='display:none' />"
            f"</body></html>"
        )
        msg.attach(MIMEText(html_full, "html", "utf-8"))

        if smtp_server:
            smtp_server.send_message(msg)
        else:
            context = ssl.create_default_context()
            with smtplib.SMTP(cfg["host"], cfg["port"], timeout=15) as server:
                server.starttls(context=context)
                server.login(cfg["user"], cfg["password"])
                server.send_message(msg)

        log = EmailLog(prospect_id=prospect.id, email_num=email_num,
                       subject=subject, body=body, status="sent",
                       tracking_id=tracking_id)
        db.session.add(log)
        prospect.emails_sent = max(prospect.emails_sent or 0, email_num)
        prospect.last_email_date = datetime.now(timezone.utc)
        prospect.date_contact_email = prospect.date_contact_email or datetime.now(timezone.utc)
        if prospect.status == "new":
            prospect.status = "email_sent"
        _calculate_score(prospect)
        return True, None

    except smtplib.SMTPRecipientsRefused:
        prospect.bounce_count = (prospect.bounce_count or 0) + 1
        prospect.status = "bounced"
        _calculate_score(prospect)
        log = EmailLog(prospect_id=prospect.id, email_num=email_num,
                       subject=subject, body=body, status="bounced",
                       error_message="Adresse email refusee", tracking_id=tracking_id)
        db.session.add(log)
        return False, "bounce"

    except smtplib.SMTPAuthenticationError as e:
        return False, f"smtp_auth:{e}"

    except Exception as e:
        log = EmailLog(prospect_id=prospect.id, email_num=email_num,
                       subject=subject, body=body, status="error",
                       error_message=str(e)[:500], tracking_id=tracking_id)
        db.session.add(log)
        return False, str(e)[:200]

# --- IMAP Helpers -------------------------------------------------------------
def _check_inbox_internal():
    """Check inbox for replies, STOP, bounces. Returns results dict."""
    imap_host = Setting.get("imap_host", "")
    imap_port = int(Setting.get("imap_port", "993"))
    imap_user = Setting.get("imap_user", Setting.get("smtp_user", ""))
    imap_pass = Setting.get("imap_pass", Setting.get("smtp_pass", ""))

    if not imap_host or not imap_user:
        return {"replies": 0, "stops": 0, "bounces": 0, "errors": 0}

    results = {"replies": 0, "stops": 0, "bounces": 0, "errors": 0}

    try:
        mail = imaplib.IMAP4_SSL(imap_host, imap_port)
        mail.login(imap_user, imap_pass)
        mail.select("INBOX")
        _, message_ids = mail.search(None, "UNSEEN")

        for mid in message_ids[0].split():
            if not mid:
                continue
            try:
                _, msg_data = mail.fetch(mid, "(RFC822)")
                msg = emaillib.message_from_bytes(msg_data[0][1])
                from_addr = emaillib.utils.parseaddr(msg.get("From", ""))[1].lower()
                subj = str(emaillib.header.decode_header(msg.get("Subject", ""))[0][0] or "")
                if isinstance(subj, bytes):
                    subj = subj.decode("utf-8", errors="ignore")

                body_text = ""
                if msg.is_multipart():
                    for part in msg.walk():
                        if part.get_content_type() == "text/plain":
                            payload = part.get_payload(decode=True)
                            if payload:
                                body_text = payload.decode("utf-8", errors="ignore")
                            break
                else:
                    payload = msg.get_payload(decode=True)
                    if payload:
                        body_text = payload.decode("utf-8", errors="ignore")

                # Detect bounce
                is_bounce = from_addr.startswith("mailer-daemon@") or from_addr.startswith("postmaster@")
                is_bounce = is_bounce or any(w in subj.lower() for w in
                    ["undeliverable", "delivery failed", "returned mail", "non remis", "undelivered"])

                if is_bounce:
                    for prospect in Prospect.query.filter(
                        Prospect.email != "", Prospect.email.isnot(None),
                        Prospect.status.notin_(["bounced", "unsubscribed"])
                    ).all():
                        if prospect.email.lower() in body_text.lower():
                            prospect.status = "bounced"
                            prospect.bounce_count = (prospect.bounce_count or 0) + 1
                            results["bounces"] += 1
                            break
                    db.session.commit()
                    continue

                # Find matching prospect
                prospect = Prospect.query.filter(db.func.lower(Prospect.email) == from_addr).first()
                if not prospect:
                    continue

                if re.search(r'\bSTOP\b', body_text, re.IGNORECASE):
                    prospect.status = "unsubscribed"
                    results["stops"] += 1
                else:
                    if prospect.status in ("email_sent", "linkedin", "both_sent"):
                        prospect.status = "replied"
                    results["replies"] += 1

                log = EmailLog.query.filter_by(prospect_id=prospect.id).order_by(EmailLog.sent_at.desc()).first()
                if log:
                    log.status = "replied"
                    log.reply_body = body_text[:2000]
                    log.reply_at = datetime.now(timezone.utc)

                _calculate_score(prospect)
                db.session.commit()
            except Exception:
                results["errors"] += 1

        mail.logout()
    except Exception as e:
        results["error"] = str(e)

    return results

# --- API: Auth ----------------------------------------------------------------
@app.route("/api/login", methods=["POST"])
@limiter.limit("5 per minute")
def login():
    data = request.get_json() or {}
    if data.get("password") == CRM_PASSWORD:
        token = _create_token()
        return jsonify({"token": token, "ok": True})
    return jsonify({"error": "Mot de passe incorrect"}), 401

# --- API: Prospects -----------------------------------------------------------
@app.route("/api/prospects")
@require_auth
def get_prospects():
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 50, type=int)
    search = request.args.get("search", "").strip()
    status = request.args.get("status", "")
    type_ = request.args.get("type", "")
    ville = request.args.get("ville", "")
    has_email = request.args.get("has_email", "")

    q = Prospect.query

    if search:
        like = f"%{search}%"
        q = q.filter(db.or_(
            Prospect.nom.ilike(like), Prospect.ville.ilike(like),
            Prospect.email.ilike(like), Prospect.region.ilike(like),
        ))
    if status:
        q = q.filter_by(status=status)
    if type_:
        q = q.filter_by(type=type_)
    if ville:
        q = q.filter_by(ville=ville)
    if has_email == "with":
        q = q.filter(Prospect.email != "", Prospect.email.isnot(None))
    elif has_email == "without":
        q = q.filter(db.or_(Prospect.email == "", Prospect.email.is_(None)))

    sort = request.args.get("sort", "date")
    if sort == "score":
        q = q.order_by(Prospect.score.desc())
    else:
        q = q.order_by(Prospect.date_ajout.desc())
    total = q.count()
    prospects = q.offset((page - 1) * per_page).limit(per_page).all()

    return jsonify({
        "prospects": [p.to_dict() for p in prospects],
        "total": total, "page": page,
        "pages": (total + per_page - 1) // per_page,
    })

@app.route("/api/prospects/stats")
@require_auth
def get_stats():
    total = Prospect.query.count()
    with_email = Prospect.query.filter(Prospect.email != "", Prospect.email.isnot(None)).count()

    status_counts = {}
    for row in db.session.query(Prospect.status, db.func.count(Prospect.id)).group_by(Prospect.status).all():
        status_counts[row[0]] = row[1]

    type_counts = {}
    for row in db.session.query(Prospect.type, db.func.count(Prospect.id)).group_by(Prospect.type).order_by(db.func.count(Prospect.id).desc()).all():
        type_counts[row[0]] = row[1]

    ville_counts = {}
    for row in db.session.query(Prospect.ville, db.func.count(Prospect.id)).group_by(Prospect.ville).order_by(db.func.count(Prospect.id).desc()).limit(20).all():
        ville_counts[row[0]] = row[1]

    email_by_type = {}
    for t, c in type_counts.items():
        with_e = Prospect.query.filter(Prospect.type == t, Prospect.email != "", Prospect.email.isnot(None)).count()
        email_by_type[t] = {"total": c, "withEmail": with_e, "pct": round(with_e / c * 100) if c > 0 else 0}

    contacted = total - status_counts.get("new", 0)
    replied = sum(status_counts.get(s, 0) for s in ["replied", "meeting", "converted"])

    return jsonify({
        "total": total, "withEmail": with_email, "contacted": contacted,
        "replied": replied, "meetings": status_counts.get("meeting", 0),
        "converted": status_counts.get("converted", 0),
        "statusCounts": status_counts, "typeCounts": type_counts,
        "villeCounts": ville_counts, "emailByType": email_by_type,
    })

@app.route("/api/prospects/<int:pid>")
@require_auth
def get_prospect(pid):
    p = Prospect.query.get_or_404(pid)
    data = p.to_dict()
    data["emailLogs"] = [l.to_dict() for l in p.email_logs.order_by(EmailLog.sent_at.desc()).all()]
    return jsonify(data)

@app.route("/api/prospects/<int:pid>", methods=["PATCH"])
@require_auth
def update_prospect(pid):
    p = Prospect.query.get_or_404(pid)
    data = request.get_json() or {}
    field_map = {
        "nom": "nom", "type": "type", "ville": "ville", "region": "region",
        "email": "email", "telephone": "telephone", "site": "site_web",
        "notes": "notes", "status": "status", "adresse": "adresse",
        "linkedinUrl": "linkedin_url",
    }
    for json_key, db_key in field_map.items():
        if json_key in data:
            setattr(p, db_key, data[json_key])
    db.session.commit()
    return jsonify(p.to_dict())

@app.route("/api/prospects", methods=["POST"])
@require_auth
def create_prospect():
    data = request.get_json() or {}
    if not data.get("nom"):
        return jsonify({"error": "Nom requis"}), 400
    p = Prospect(
        nom=data["nom"], type=data.get("type", "hebergement"),
        ville=data.get("ville", ""), region=data.get("region", data.get("ville", "")),
        email=data.get("email", ""), telephone=data.get("telephone", ""),
        site_web=data.get("site", ""), notes=data.get("notes", ""), status="new",
        unsubscribe_token=secrets.token_urlsafe(32),
    )
    _calculate_score(p)
    db.session.add(p)
    db.session.commit()
    return jsonify(p.to_dict()), 201

@app.route("/api/prospects/<int:pid>", methods=["DELETE"])
@require_auth
def delete_prospect(pid):
    p = Prospect.query.get_or_404(pid)
    EmailLog.query.filter_by(prospect_id=pid).delete()
    db.session.delete(p)
    db.session.commit()
    return jsonify({"ok": True})

# --- API: Filters -------------------------------------------------------------
@app.route("/api/filters")
@require_auth
def get_filters():
    types = [r[0] for r in db.session.query(Prospect.type).distinct().order_by(Prospect.type).all() if r[0]]
    villes = [r[0] for r in db.session.query(Prospect.ville, db.func.count(Prospect.id)).group_by(Prospect.ville).order_by(db.func.count(Prospect.id).desc()).limit(50).all() if r[0]]
    return jsonify({"types": types, "villes": villes})

# --- API: Send single email (manual) -----------------------------------------
@app.route("/api/send-email", methods=["POST"])
@require_auth
def send_email():
    data = request.get_json() or {}
    pid = data.get("prospectId")
    subject = data.get("subject", "")
    body = data.get("body", "")
    email_num = data.get("emailNum", 1)

    p = Prospect.query.get_or_404(pid)
    if not p.email:
        return jsonify({"error": "Ce prospect n'a pas d'adresse email"}), 400

    cfg = _get_smtp_config()
    if not cfg["host"] or not cfg["user"]:
        return jsonify({"error": "SMTP non configure. Allez dans Parametres."}), 400

    success, error = _send_one_email(p, subject, body, email_num)
    db.session.commit()

    if success:
        return jsonify({"ok": True, "message": f"Email envoye a {p.email}"})
    if error and error.startswith("smtp_auth"):
        return jsonify({"error": "Erreur authentification SMTP"}), 500
    return jsonify({"error": f"Erreur envoi: {error}"}), 500

# --- API: Check Inbox (IMAP) -------------------------------------------------
@app.route("/api/check-inbox", methods=["POST"])
@require_auth
def check_inbox():
    results = _check_inbox_internal()
    if "error" in results and isinstance(results["error"], str) and "configure" in results["error"].lower():
        return jsonify({"error": results["error"]}), 400
    return jsonify({"ok": True, **results})

# --- API: Auto-Send Campaign Engine -------------------------------------------
@app.route("/api/auto-send", methods=["POST"])
@require_auth_or_secret
@limiter.limit("2 per minute")
def auto_send():
    """The main campaign engine. Sends email sequences automatically."""
    cfg = _get_smtp_config()
    if not cfg["host"] or not cfg["user"]:
        return jsonify({"error": "SMTP non configure"}), 400

    # Smart send hours check
    send_start = int(Setting.get("send_hour_start", "9"))
    send_end = int(Setting.get("send_hour_end", "18"))
    france_hour = _get_france_hour()
    if not (send_start <= france_hour < send_end):
        return jsonify({"ok": True, "sent": 0, "failed": 0, "outsideHours": True,
                        "message": f"Hors heures d'envoi ({send_start}h-{send_end}h France). Il est {france_hour}h."})

    daily_limit = int(Setting.get("daily_limit", "30"))
    delay_email2 = int(Setting.get("delay_email2", "3"))
    delay_email3 = int(Setting.get("delay_email3", "5"))
    max_per_run = int(request.args.get("max", "15"))

    # Count emails sent today
    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    sent_today = EmailLog.query.filter(
        EmailLog.sent_at >= today_start, EmailLog.status == "sent"
    ).count()

    remaining = daily_limit - sent_today
    if remaining <= 0:
        return jsonify({"ok": True, "message": "Limite journaliere atteinte",
                        "sentToday": sent_today, "dailyLimit": daily_limit,
                        "sent": 0, "failed": 0})

    # Check inbox first to update statuses
    inbox_results = _check_inbox_internal()

    # Build send queue: Email 1 > Email 2 > Email 3
    queue = []
    now = datetime.now(timezone.utc)

    # Email 1: new prospects with email + template
    email1_candidates = Prospect.query.filter(
        Prospect.status == "new",
        Prospect.email != "", Prospect.email.isnot(None),
        Prospect.email1_sujet != "", Prospect.email1_sujet.isnot(None),
        Prospect.email1_corps != "", Prospect.email1_corps.isnot(None),
        Prospect.emails_sent == 0,
    ).order_by(Prospect.date_ajout.asc()).all()

    for p in email1_candidates:
        if _is_valid_email(p.email):
            queue.append((p, 1, p.email1_sujet, p.email1_corps))

    # Email 2: sent email 1, waited delay_email2 days
    cutoff2 = now - timedelta(days=delay_email2)
    email2_candidates = Prospect.query.filter(
        Prospect.emails_sent == 1,
        Prospect.last_email_date <= cutoff2,
        Prospect.status.notin_(SKIP_STATUSES),
        Prospect.email2_sujet != "", Prospect.email2_sujet.isnot(None),
        Prospect.email2_corps != "", Prospect.email2_corps.isnot(None),
    ).order_by(Prospect.last_email_date.asc()).all()

    for p in email2_candidates:
        if _is_valid_email(p.email):
            queue.append((p, 2, p.email2_sujet, p.email2_corps))

    # Email 3: sent email 2, waited delay_email3 days
    cutoff3 = now - timedelta(days=delay_email3)
    email3_candidates = Prospect.query.filter(
        Prospect.emails_sent == 2,
        Prospect.last_email_date <= cutoff3,
        Prospect.status.notin_(SKIP_STATUSES),
        Prospect.email3_sujet != "", Prospect.email3_sujet.isnot(None),
        Prospect.email3_corps != "", Prospect.email3_corps.isnot(None),
    ).order_by(Prospect.last_email_date.asc()).all()

    for p in email3_candidates:
        if _is_valid_email(p.email):
            queue.append((p, 3, p.email3_sujet, p.email3_corps))

    # Truncate to limits
    queue = queue[:min(remaining, max_per_run)]

    if not queue:
        return jsonify({"ok": True, "message": "Aucun email a envoyer",
                        "sent": 0, "failed": 0, "sentToday": sent_today,
                        "dailyLimit": daily_limit, "inbox": inbox_results,
                        "queueSize": 0})

    # Create campaign run
    run = CampaignRun(status="running", **inbox_results_to_run(inbox_results))
    db.session.add(run)
    db.session.commit()

    # Open SMTP connection once, deduplicate emails
    results = {"sent": 0, "failed": 0, "bounced": 0, "duplicates": 0, "details": []}
    seen_emails = set()
    smtp_server = None
    try:
        smtp_server = _open_smtp_connection()
    except Exception as e:
        run.status = "error"
        run.details = json.dumps({"error": f"SMTP connection failed: {str(e)}"})
        run.finished_at = datetime.now(timezone.utc)
        db.session.commit()
        return jsonify({"error": f"Connexion SMTP impossible: {str(e)}"}), 500

    for i, (prospect, email_num, subject, body) in enumerate(queue):
        success, error = _send_one_email(prospect, subject, body, email_num, smtp_server, seen_emails)
        db.session.commit()

        detail = {"prospectId": prospect.id, "nom": prospect.nom,
                  "emailNum": email_num, "success": success}
        if error:
            detail["error"] = error

        results["details"].append(detail)

        if success:
            results["sent"] += 1
        else:
            results["failed"] += 1
            if error == "bounce":
                results["bounced"] += 1
            elif error == "duplicate":
                results["duplicates"] += 1
            elif error and error.startswith("smtp_auth"):
                # Auth failure = stop campaign
                break

        # Delay between sends (2-5 seconds)
        if i < len(queue) - 1:
            time.sleep(random.uniform(2, 4))

    # Close SMTP
    if smtp_server:
        try:
            smtp_server.quit()
        except Exception:
            pass

    # Update campaign run
    run.emails_sent = results["sent"]
    run.emails_failed = results["failed"]
    run.bounces_detected += results["bounced"]
    run.status = "completed"
    run.finished_at = datetime.now(timezone.utc)
    run.details = json.dumps(results["details"])
    db.session.commit()

    return jsonify({
        "ok": True, "sent": results["sent"], "failed": results["failed"],
        "bounced": results["bounced"], "sentToday": sent_today + results["sent"],
        "dailyLimit": daily_limit, "inbox": inbox_results,
        "details": results["details"],
    })

def inbox_results_to_run(inbox):
    """Convert inbox check results to CampaignRun fields."""
    return {
        "replies_detected": inbox.get("replies", 0),
        "stops_detected": inbox.get("stops", 0),
        "bounces_detected": inbox.get("bounces", 0),
    }

# --- API: Campaign Stats -----------------------------------------------------
@app.route("/api/campaign/stats")
@require_auth
def campaign_stats():
    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    sent_today = EmailLog.query.filter(
        EmailLog.sent_at >= today_start, EmailLog.status == "sent"
    ).count()
    daily_limit = int(Setting.get("daily_limit", "30"))

    # Ready to send counts
    ready_email1 = Prospect.query.filter(
        Prospect.status == "new", Prospect.email != "", Prospect.email.isnot(None),
        Prospect.email1_sujet != "", Prospect.email1_sujet.isnot(None),
        Prospect.emails_sent == 0,
    ).count()

    delay2 = int(Setting.get("delay_email2", "3"))
    delay3 = int(Setting.get("delay_email3", "5"))
    now = datetime.now(timezone.utc)

    ready_email2 = Prospect.query.filter(
        Prospect.emails_sent == 1,
        Prospect.last_email_date <= now - timedelta(days=delay2),
        Prospect.status.notin_(SKIP_STATUSES),
        Prospect.email2_sujet != "", Prospect.email2_sujet.isnot(None),
    ).count()

    ready_email3 = Prospect.query.filter(
        Prospect.emails_sent == 2,
        Prospect.last_email_date <= now - timedelta(days=delay3),
        Prospect.status.notin_(SKIP_STATUSES),
        Prospect.email3_sujet != "", Prospect.email3_sujet.isnot(None),
    ).count()

    # Sequence progress
    at_stage0 = Prospect.query.filter(Prospect.emails_sent == 0, Prospect.email != "", Prospect.email.isnot(None)).count()
    at_stage1 = Prospect.query.filter(Prospect.emails_sent == 1).count()
    at_stage2 = Prospect.query.filter(Prospect.emails_sent == 2).count()
    at_stage3 = Prospect.query.filter(Prospect.emails_sent >= 3).count()

    # Last run
    last_run = CampaignRun.query.order_by(CampaignRun.started_at.desc()).first()

    return jsonify({
        "sentToday": sent_today, "dailyLimit": daily_limit,
        "readyEmail1": ready_email1, "readyEmail2": ready_email2, "readyEmail3": ready_email3,
        "atStage0": at_stage0, "atStage1": at_stage1, "atStage2": at_stage2, "atStage3": at_stage3,
        "lastRun": last_run.to_dict() if last_run else None,
    })

# --- API: Bulk generate templates ---------------------------------------------
@app.route("/api/bulk-generate-emails", methods=["POST"])
@require_auth
def bulk_generate_emails():
    """Apply email templates to prospects that don't have them yet."""
    data = request.get_json() or {}
    templates = {
        "email1_sujet": data.get("email1Sujet", ""),
        "email1_corps": data.get("email1Corps", ""),
        "email2_sujet": data.get("email2Sujet", ""),
        "email2_corps": data.get("email2Corps", ""),
        "email3_sujet": data.get("email3Sujet", ""),
        "email3_corps": data.get("email3Corps", ""),
    }

    # Target: prospects without templates that have an email
    prospects = Prospect.query.filter(
        Prospect.email != "", Prospect.email.isnot(None),
        db.or_(Prospect.email1_sujet == "", Prospect.email1_sujet.is_(None)),
    ).all()

    count = 0
    for p in prospects:
        for field, value in templates.items():
            if value:
                setattr(p, field, value)
        count += 1

    db.session.commit()
    return jsonify({"ok": True, "updated": count})

# --- API: Settings ------------------------------------------------------------
@app.route("/api/settings", methods=["GET"])
@require_auth
def get_settings():
    keys = ["smtp_host", "smtp_port", "smtp_user", "smtp_pass", "sender_email",
            "sender_name", "imap_host", "imap_port", "imap_user", "imap_pass",
            "daily_limit", "delay_email2", "delay_email3",
            "send_hour_start", "send_hour_end", "base_url",
            "tpl_email1_sujet", "tpl_email1_corps",
            "tpl_email2_sujet", "tpl_email2_corps",
            "tpl_email3_sujet", "tpl_email3_corps"]
    return jsonify({k: Setting.get(k, "") for k in keys})

@app.route("/api/settings", methods=["POST"])
@require_auth
def save_settings():
    data = request.get_json() or {}
    for key, value in data.items():
        Setting.set(key, str(value))
    return jsonify({"ok": True})

# --- API: Bulk Actions --------------------------------------------------------
@app.route("/api/bulk-status", methods=["POST"])
@require_auth
def bulk_status():
    data = request.get_json() or {}
    ids = data.get("ids", [])
    new_status = data.get("status", "")
    if not ids or not new_status:
        return jsonify({"error": "ids et status requis"}), 400
    Prospect.query.filter(Prospect.id.in_(ids)).update({Prospect.status: new_status}, synchronize_session=False)
    db.session.commit()
    return jsonify({"ok": True, "updated": len(ids)})

# --- API: Import (smart dedup) ------------------------------------------------
@app.route("/api/import", methods=["POST"])
@require_auth
def import_prospects():
    data = request.get_json() or {}
    prospects_data = data.get("prospects", [])
    stats = {"imported": 0, "skipped": 0, "updated": 0, "errors": 0}

    for pd_item in prospects_data:
        nom = (pd_item.get("nom") or "").strip()
        if not nom:
            stats["errors"] += 1
            continue

        email_addr = (pd_item.get("email") or "").strip().lower()
        ville = (pd_item.get("ville") or "").strip()

        # Deduplication: by email OR nom+ville
        existing = None
        if email_addr:
            existing = Prospect.query.filter(db.func.lower(Prospect.email) == email_addr).first()
        if not existing and nom and ville:
            existing = Prospect.query.filter(
                db.func.lower(Prospect.nom) == nom.lower(),
                db.func.lower(Prospect.ville) == ville.lower()
            ).first()

        if existing:
            updated = False
            for field, attr in [("email", "email"), ("telephone", "telephone"),
                                ("site", "site_web"), ("adresse", "adresse"),
                                ("note", "note_google"), ("avis", "nb_avis"),
                                ("googleMaps", "google_maps")]:
                new_val = pd_item.get(field)
                if new_val and not getattr(existing, attr):
                    setattr(existing, attr, new_val)
                    updated = True
            stats["updated" if updated else "skipped"] += 1
            continue

        p = Prospect(
            nom=nom, type=pd_item.get("type", ""), ville=ville,
            region=pd_item.get("region", ville),
            email=email_addr, telephone=pd_item.get("telephone", ""),
            site_web=pd_item.get("site", ""), note_google=pd_item.get("note", 0),
            nb_avis=pd_item.get("avis", 0), status="new",
            adresse=pd_item.get("adresse", ""), google_maps=pd_item.get("googleMaps", ""),
        )
        db.session.add(p)
        stats["imported"] += 1

    db.session.commit()
    return jsonify({"ok": True, **stats})

# --- API: Open Tracking (pixel) -----------------------------------------------
@app.route("/api/track/<tracking_id>")
def track_open(tracking_id):
    """Track email open via 1x1 pixel. No auth required."""
    from flask import Response
    log = EmailLog.query.filter_by(tracking_id=tracking_id).first()
    if log:
        log.open_count = (log.open_count or 0) + 1
        if not log.opened_at:
            log.opened_at = datetime.now(timezone.utc)
        prospect = Prospect.query.get(log.prospect_id)
        if prospect:
            prospect.email_opened = True
            _calculate_score(prospect)
        db.session.commit()
    return Response(TRACKING_GIF, mimetype="image/gif",
                    headers={"Cache-Control": "no-cache, no-store, must-revalidate"})

# --- API: Unsubscribe ---------------------------------------------------------
@app.route("/api/unsubscribe/<token>")
def unsubscribe(token):
    """Public unsubscribe link. No auth required."""
    prospect = Prospect.query.filter_by(unsubscribe_token=token).first()
    if prospect:
        prospect.status = "unsubscribed"
        _calculate_score(prospect)
        db.session.commit()
    html = """<!DOCTYPE html><html><head><meta charset="utf-8"><title>Desinscription</title>
    <style>body{font-family:Arial,sans-serif;display:flex;align-items:center;justify-content:center;
    min-height:100vh;background:#f9fafb;margin:0;}
    .card{background:white;border-radius:16px;padding:40px;text-align:center;box-shadow:0 4px 12px rgba(0,0,0,0.1);max-width:400px;}
    </style></head><body><div class="card">
    <h2 style="color:#059669;">Desinscription confirmee</h2>
    <p style="color:#6B7280;">Vous ne recevrez plus d'emails de notre part.</p>
    </div></body></html>"""
    return html, 200

# --- API: Recalculate Scores --------------------------------------------------
@app.route("/api/recalculate-scores", methods=["POST"])
@require_auth
def recalculate_scores():
    """Recalculate all prospect scores."""
    count = 0
    for p in Prospect.query.all():
        _calculate_score(p)
        count += 1
    db.session.commit()
    return jsonify({"ok": True, "updated": count})

# --- API: Email Analytics -----------------------------------------------------
@app.route("/api/analytics/emails")
@require_auth
def email_analytics():
    """Per-template analytics: open rate, reply rate, bounce rate."""
    analytics = {}
    for num in [1, 2, 3]:
        logs = EmailLog.query.filter_by(email_num=num)
        total = logs.count()
        if total == 0:
            analytics[f"email{num}"] = {"total": 0, "opened": 0, "replied": 0, "bounced": 0,
                                         "openRate": 0, "replyRate": 0, "bounceRate": 0}
            continue
        opened = logs.filter(EmailLog.opened_at.isnot(None)).count()
        replied = logs.filter(EmailLog.status == "replied").count()
        bounced = logs.filter(EmailLog.status == "bounced").count()
        analytics[f"email{num}"] = {
            "total": total, "opened": opened, "replied": replied, "bounced": bounced,
            "openRate": round(opened / total * 100, 1) if total else 0,
            "replyRate": round(replied / total * 100, 1) if total else 0,
            "bounceRate": round(bounced / total * 100, 1) if total else 0,
        }

    # Global funnel from prospect statuses
    funnel = {}
    for s in ["new", "email_sent", "replied", "meeting", "converted"]:
        funnel[s] = Prospect.query.filter_by(status=s).count()
    analytics["funnel"] = funnel

    # Hot leads
    analytics["hotLeads"] = Prospect.query.filter(Prospect.score >= 30).count()

    return jsonify(analytics)

# --- API: Preview Email -------------------------------------------------------
@app.route("/api/preview-email", methods=["POST"])
@require_auth
def preview_email():
    """Preview a template rendered with a prospect's data."""
    data = request.get_json() or {}
    pid = data.get("prospectId")
    subject = data.get("subject", "")
    body = data.get("body", "")

    if pid:
        p = Prospect.query.get_or_404(pid)
    else:
        # Pick a random prospect for preview
        p = Prospect.query.filter(Prospect.email != "", Prospect.email.isnot(None)).first()
        if not p:
            return jsonify({"error": "Aucun prospect avec email"}), 404

    rendered_subject = _render_template(subject, p)
    rendered_body = _render_template(body, p)
    html_body = rendered_body.replace("\n", "<br>")
    return jsonify({
        "subject": rendered_subject,
        "bodyText": rendered_body,
        "bodyHtml": f"<div style='font-family:Arial;font-size:14px;color:#333;line-height:1.6'>{html_body}</div>",
        "prospect": {"nom": p.nom, "email": p.email, "ville": p.ville},
    })

# --- API: Export CSV ----------------------------------------------------------
@app.route("/api/export")
@require_auth
def export_prospects():
    """Export filtered prospects to CSV."""
    import csv
    import io
    from flask import Response

    q = Prospect.query
    status = request.args.get("status", "")
    type_ = request.args.get("type", "")
    search = request.args.get("search", "").strip()

    if search:
        like = f"%{search}%"
        q = q.filter(db.or_(Prospect.nom.ilike(like), Prospect.ville.ilike(like),
                             Prospect.email.ilike(like)))
    if status:
        q = q.filter_by(status=status)
    if type_:
        q = q.filter_by(type=type_)

    prospects = q.order_by(Prospect.score.desc()).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Nom", "Type", "Ville", "Region", "Email", "Telephone",
                     "Site Web", "Note Google", "Avis", "Statut", "Score",
                     "Emails Envoyes", "Email Ouvert", "Date Ajout"])
    for p in prospects:
        writer.writerow([p.nom, p.type, p.ville, p.region, p.email, p.telephone,
                         p.site_web, p.note_google, p.nb_avis, p.status,
                         p.score or 0, p.emails_sent or 0,
                         "Oui" if p.email_opened else "Non",
                         p.date_ajout.strftime("%Y-%m-%d") if p.date_ajout else ""])

    return Response(output.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": "attachment; filename=prospects_jalunia.csv"})

# --- Health Check -------------------------------------------------------------
@app.route("/health")
def health():
    try:
        db.session.execute(db.text("SELECT 1"))
        return jsonify({"status": "healthy", "database": "connected"}), 200
    except Exception as e:
        return jsonify({"status": "unhealthy", "database": str(e)}), 503

@app.route("/debug")
def debug():
    """Diagnostic endpoint - shows database state and any errors."""
    info = {}
    try:
        info["db"] = "connected"
        # Check if tables exist
        result = db.session.execute(db.text(
            "SELECT column_name FROM information_schema.columns WHERE table_name='prospects' ORDER BY ordinal_position"
        ))
        info["prospect_columns"] = [r[0] for r in result]
    except Exception as e:
        info["db_error"] = str(e)
        # Try SQLite fallback
        try:
            result = db.session.execute(db.text("PRAGMA table_info(prospects)"))
            info["prospect_columns"] = [r[1] for r in result]
        except Exception as e2:
            info["db_error_sqlite"] = str(e2)
    try:
        count = Prospect.query.count()
        info["prospect_count"] = count
    except Exception as e:
        info["prospect_query_error"] = str(e)
    try:
        p = Prospect.query.first()
        if p:
            info["first_prospect"] = p.nom
            info["first_prospect_dict"] = p.to_dict()
        else:
            info["first_prospect"] = "NONE - table is empty"
    except Exception as e:
        info["to_dict_error"] = str(e)
    try:
        log_count = EmailLog.query.count()
        info["email_log_count"] = log_count
    except Exception as e:
        info["email_log_error"] = str(e)
    return jsonify(info)

# --- Serve Frontend -----------------------------------------------------------
@app.route("/")
def index():
    return send_from_directory("static", "index.html")

@app.route("/<path:path>")
def static_files(path):
    return send_from_directory("static", path)

# --- Init Script: Import from prospects.xlsx ----------------------------------
def import_from_excel(xlsx_path):
    import openpyxl

    STATUS_MAP = {
        "new": "new", "email_sent": "email_sent",
        "linkedin": "linkedin", "both_sent": "both_sent",
        "replied": "replied", "meeting": "meeting",
        "converted": "converted", "not_interested": "not_interested",
        "follow_up": "follow_up", "unsubscribed": "unsubscribed",
    }

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    imported = 0

    for r in range(2, ws.max_row + 1):
        nom = ws.cell(r, 1).value
        if not nom:
            continue
        nom = str(nom).strip()

        existing = Prospect.query.filter_by(nom=nom).first()
        if existing:
            continue

        p = Prospect(
            nom=nom,
            type=str(ws.cell(r, 2).value or "").strip(),
            ville=str(ws.cell(r, 3).value or "").strip(),
            region=str(ws.cell(r, 4).value or "").strip(),
            adresse=str(ws.cell(r, 5).value or "").strip(),
            telephone=str(ws.cell(r, 6).value or "").strip(),
            email=str(ws.cell(r, 7).value or "").strip(),
            site_web=str(ws.cell(r, 8).value or "").strip(),
            google_maps=str(ws.cell(r, 9).value or "").strip(),
            note_google=float(ws.cell(r, 10).value or 0),
            nb_avis=int(ws.cell(r, 11).value or 0),
            status="new",
        )
        db.session.add(p)
        imported += 1

        if imported % 500 == 0:
            db.session.commit()

    db.session.commit()
    print(f"Import termine: {imported} prospects importes")


if __name__ == "__main__":
    import sys

    if len(sys.argv) > 1 and sys.argv[1] == "import":
        xlsx_path = sys.argv[2] if len(sys.argv) > 2 else "../prospecting/prospects.xlsx"
        with app.app_context():
            print(f"Import depuis {xlsx_path}...")
            import_from_excel(xlsx_path)
    else:
        port = int(os.environ.get("PORT", 5000))
        app.run(host="0.0.0.0", port=port, debug=os.environ.get("FLASK_DEBUG", "0") == "1")
